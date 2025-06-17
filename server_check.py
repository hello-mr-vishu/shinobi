import os
import datetime
import ping3
import openpyxl
import schedule
import time
import subprocess
import platform
import requests
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Configuration: Your servers list
servers = [
    {"KETL": "KETL", "ip": "172.24.24.146"},
    {"SMTL": "SMTL", "ip": "172.24.179.113"},
    {"NKTL": "NKTL", "ip": "172.24.134.79"},
    {"LRTL": "LRTL", "ip": "172.24.61.249"},
    {"JVTL": "JVTL", "ip": "172.24.48.115"},
    {"MKTL-15": "MKTL-15", "ip": "172.24.15.161"},
    {"MKTL-16": "MKTL-16", "ip": "172.24.137.166"},
    {"JLTL": "JLTL", "ip": "172.24.116.111"},
]

# Directory to save Excel files
LOG_DIR = "server_logs"
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)

# Track server state
server_status_tracker = {
    server[next(iter(server))]: {
        "last_status": None,
        "last_down_time": None,
        "last_up_time": None,
        "notification_sent": False
    } for server in servers
}

def get_excel_filename():
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
    return os.path.join(LOG_DIR, f"server_status_{current_date}.xlsx")

def ping_server(ip):
    def try_ping():
        try:
            response_time = ping3.ping(ip, timeout=5)
            if response_time is not None:
                print(f"Ping3 success for {ip}: {response_time}ms")
                return True
        except Exception as e:
            print(f"Ping3 error for {ip}: {e}")
        try:
            param = '-n' if platform.system().lower() == 'windows' else '-c'
            result = subprocess.run(['ping', param, '1', ip], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=10)
            return result.returncode == 0
        except Exception as e:
            print(f"System ping error for {ip}: {e}")
            return False

    if try_ping():
        return "Up"
    print(f"Initial ping failed for {ip}, waiting 120 seconds...")
    time.sleep(120)
    if try_ping():
        return "Up"
    return "Down"

def initialize_excel_sheet(workbook, sheet_name):
    sheet_name = sheet_name.replace("/", "_").replace("\\", "_").replace(":", "_")
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        headers = ["Date & Time", "Server IP", "Server Status", "Last Down Time", "Last Up Time", "Downtime Duration (Min)"]
        for col_num, header in enumerate(headers, 1):
            sheet[f"{get_column_letter(col_num)}1"] = header
    return workbook[sheet_name]

def send_whatsapp_alert(number, message):
    try:
        res = requests.post("http://localhost:3000/send", json={
            "number": number,
            "message": message
        })
        if res.status_code == 200:
            print("✅ WhatsApp alert sent.")
        else:
            print(f"❌ WhatsApp alert failed: {res.text}")
    except Exception as e:
        print(f"❌ Error sending WhatsApp alert: {e}")

def log_server_status():
    print("Servers configuration:", servers)
    excel_file = get_excel_filename()

    try:
        workbook = openpyxl.load_workbook(excel_file)
    except Exception:
        print(f"Creating new workbook: {excel_file}")
        workbook = openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    data_logged = False

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for server in servers:
        server_key = next(iter(server))
        server_name = server[server_key]
        ip = server["ip"]
        status = ping_server(ip)
        down_time = ""
        up_time = ""
        duration = ""

        current_time_dt = datetime.datetime.strptime(current_time, "%Y-%m-%d %H:%M:%S")
        tracker = server_status_tracker[server_key]
        last_status = tracker["last_status"]

        if status == "Down" and not tracker["notification_sent"]:
            message = f"🚨 Server Down Alert 🚨\nServer: {server_name}\nIP: {ip}\nTime: {current_time}"
            send_whatsapp_alert("8328618110", message)  # Replace with real number
            tracker["notification_sent"] = True

        elif status == "Up":
            if last_status == "Down":
                tracker["last_up_time"] = current_time_dt
                up_time = current_time
                if tracker["last_down_time"]:
                    duration_dt = current_time_dt - tracker["last_down_time"] 
                    duration = round(duration_dt.total_seconds() / 60, 2)
                    tracker["last_down_time"] = None
                tracker["notification_sent"] = False
            if tracker["last_up_time"]:
                up_time = tracker["last_up_time"].strftime("%Y-%m-%d %H:%M:%S")

        if status == "Down" and last_status != "Down":
            tracker["last_down_time"] = current_time_dt

        tracker["last_status"] = status

        sheet = initialize_excel_sheet(workbook, server_name)
        row = [current_time, ip, status, down_time, up_time, duration]
        sheet.append(row)
        data_logged = True

        status_cell = f"C{sheet.max_row}"
        sheet[status_cell].fill = red_fill if status == "Down" else green_fill

        if status == "Down":
            print(f"Logged: {server_name} ({ip}) - Down at {current_time}")
        else:
            print(f"Logged: {server_name} ({ip}) - Up at {current_time}, Downtime: {duration} min")

    if data_logged:
        try:
            workbook.save(excel_file)
            print(f"Saved log to {excel_file}\n")
        except Exception as e:
            print(f"❌ Error saving Excel file: {e}")
    else:
        print("No new data to log.")

def main():
    schedule.every(5).minutes.do(log_server_status)
    print("🔍 Server status monitoring started...\n")
    log_server_status()

    while True:
        try:
            schedule.run_pending()
            time.sleep(1)
        except KeyboardInterrupt:
            print("⛔ Monitoring stopped by user.")
            break
        except Exception as e:
            print(f"❌ Unexpected error: {e}")
            time.sleep(60)

if __name__ == "__main__":
    main()
